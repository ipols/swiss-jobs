"""
Fetch Swiss occupation data from BFS (Federal Statistical Office).

Data sources:
- Employment: BFS SAKE 2025 (Swiss Labour Force Survey) via DAM Excel download
  Asset 36346663 — Employment by CH-ISCO-19 major groups, annual averages
- Wages: BFS ESS 2022 (Earnings Structure Survey) via STAT-TAB PxWeb API
  Table px-x-0304010000_205 — Median gross monthly wages by ISCO 2-digit group

Output: data/occupations.json
"""

import json
import os
import httpx
import openpyxl

# ISCO-08 occupation classification at 2-digit sub-major group level
# Identical to CH-ISCO-19 levels 1-2
# Source: ILO ISCO-08 structure
ISCO_2DIGIT = {
    # Major group 1: Managers
    "11": {"title": "Chief Executives, Senior Officials and Legislators", "major": "1", "major_name": "Managers", "skill_level": 4},
    "12": {"title": "Administrative and Commercial Managers", "major": "1", "major_name": "Managers", "skill_level": 4},
    "13": {"title": "Production and Specialised Services Managers", "major": "1", "major_name": "Managers", "skill_level": 4},
    "14": {"title": "Hospitality, Retail and Other Services Managers", "major": "1", "major_name": "Managers", "skill_level": 4},
    # Major group 2: Professionals
    "21": {"title": "Science and Engineering Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    "22": {"title": "Health Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    "23": {"title": "Teaching Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    "24": {"title": "Business and Administration Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    "25": {"title": "ICT Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    "26": {"title": "Legal, Social and Cultural Professionals", "major": "2", "major_name": "Professionals", "skill_level": 4},
    # Major group 3: Technicians and Associate Professionals
    "31": {"title": "Science and Engineering Associate Professionals", "major": "3", "major_name": "Technicians", "skill_level": 3},
    "32": {"title": "Health Associate Professionals", "major": "3", "major_name": "Technicians", "skill_level": 3},
    "33": {"title": "Business and Administration Associate Professionals", "major": "3", "major_name": "Technicians", "skill_level": 3},
    "34": {"title": "Legal, Social, Cultural and Related Associate Professionals", "major": "3", "major_name": "Technicians", "skill_level": 3},
    "35": {"title": "Information and Communications Technicians", "major": "3", "major_name": "Technicians", "skill_level": 3},
    # Major group 4: Clerical Support Workers
    "41": {"title": "General and Keyboard Clerks", "major": "4", "major_name": "Clerical Support", "skill_level": 2},
    "42": {"title": "Customer Services Clerks", "major": "4", "major_name": "Clerical Support", "skill_level": 2},
    "43": {"title": "Numerical and Material Recording Clerks", "major": "4", "major_name": "Clerical Support", "skill_level": 2},
    "44": {"title": "Other Clerical Support Workers", "major": "4", "major_name": "Clerical Support", "skill_level": 2},
    # Major group 5: Service and Sales Workers
    "51": {"title": "Personal Service Workers", "major": "5", "major_name": "Services & Sales", "skill_level": 2},
    "52": {"title": "Sales Workers", "major": "5", "major_name": "Services & Sales", "skill_level": 2},
    "53": {"title": "Personal Care Workers", "major": "5", "major_name": "Services & Sales", "skill_level": 2},
    "54": {"title": "Protective Services Workers", "major": "5", "major_name": "Services & Sales", "skill_level": 2},
    # Major group 6: Skilled Agricultural, Forestry and Fishery Workers
    "61": {"title": "Market-oriented Skilled Agricultural Workers", "major": "6", "major_name": "Agriculture", "skill_level": 2},
    "62": {"title": "Market-oriented Skilled Forestry, Fishery and Hunting Workers", "major": "6", "major_name": "Agriculture", "skill_level": 2},
    # Major group 7: Craft and Related Trades Workers
    "71": {"title": "Building and Related Trades Workers (excl. Electricians)", "major": "7", "major_name": "Craft & Trades", "skill_level": 2},
    "72": {"title": "Metal, Machinery and Related Trades Workers", "major": "7", "major_name": "Craft & Trades", "skill_level": 2},
    "73": {"title": "Handicraft and Printing Workers", "major": "7", "major_name": "Craft & Trades", "skill_level": 2},
    "74": {"title": "Electrical and Electronic Trades Workers", "major": "7", "major_name": "Craft & Trades", "skill_level": 2},
    "75": {"title": "Food Processing, Woodworking, Garment and Other Craft Workers", "major": "7", "major_name": "Craft & Trades", "skill_level": 2},
    # Major group 8: Plant and Machine Operators and Assemblers
    "81": {"title": "Stationary Plant and Machine Operators", "major": "8", "major_name": "Machine Operators", "skill_level": 1},
    "82": {"title": "Assemblers", "major": "8", "major_name": "Machine Operators", "skill_level": 1},
    "83": {"title": "Drivers and Mobile Plant Operators", "major": "8", "major_name": "Machine Operators", "skill_level": 2},
    # Major group 9: Elementary Occupations
    "91": {"title": "Cleaners and Helpers", "major": "9", "major_name": "Elementary", "skill_level": 1},
    "92": {"title": "Agricultural, Forestry and Fishery Labourers", "major": "9", "major_name": "Elementary", "skill_level": 1},
    "93": {"title": "Labourers in Mining, Construction, Manufacturing and Transport", "major": "9", "major_name": "Elementary", "skill_level": 1},
    "94": {"title": "Food Preparation Assistants", "major": "9", "major_name": "Elementary", "skill_level": 1},
    "96": {"title": "Refuse Workers and Other Elementary Workers", "major": "9", "major_name": "Elementary", "skill_level": 1},
}

SKILL_LEVEL_LABELS = {
    1: "No formal / primary education",
    2: "Upper secondary / apprenticeship",
    3: "Higher vocational / short-cycle tertiary",
    4: "Bachelor's degree or higher",
}

# Map BFS German major group names to ISCO 1-digit codes
BFS_MAJOR_GROUP_NAMES = {
    "Führungskräfte": "1",
    "Intellektuelle und wissenschaftliche Berufe": "2",
    "Techniker/innen und gleichrangige nichttechnische Berufe": "3",
    "Bürokräfte und verwandte Berufe": "4",
    "Dienstleistungsberufe und Verkäufer/innen": "5",
    "Fachkräfte in Land- und Forstwirtschaft und Fischerei": "6",
    "Fachkräfte in der Land- und Forstwirtschaft und Fischerei": "6",
    "Handwerks- und verwandte Berufe": "7",
    "Bediener/innen von Anlagen und Maschinen und Montageberufe": "8",
    "Hilfsarbeitskräfte": "9",
}

# Relative weights for distributing 1-digit employment to 2-digit sub-groups
# Based on European LFS proportions for Switzerland
SUBGROUP_WEIGHTS = {
    "1": {"11": 0.15, "12": 0.35, "13": 0.30, "14": 0.20},
    "2": {"21": 0.20, "22": 0.16, "23": 0.18, "24": 0.22, "25": 0.14, "26": 0.10},
    "3": {"31": 0.22, "32": 0.14, "33": 0.40, "34": 0.14, "35": 0.10},
    "4": {"41": 0.33, "42": 0.22, "43": 0.30, "44": 0.15},
    "5": {"51": 0.30, "52": 0.30, "53": 0.25, "54": 0.15},
    "6": {"61": 0.85, "62": 0.15},
    "7": {"71": 0.32, "72": 0.28, "73": 0.07, "74": 0.17, "75": 0.16},
    "8": {"81": 0.25, "82": 0.20, "83": 0.55},
    "9": {"91": 0.40, "92": 0.05, "93": 0.25, "94": 0.15, "96": 0.15},
}


def download_employment_excel():
    """Download BFS SAKE employment Excel if not cached."""
    path = "data/employment_raw.xlsx"
    if os.path.exists(path):
        print(f"  Using cached {path}")
        return path

    print("  Downloading BFS SAKE employment data (asset 36346663)...")
    r = httpx.get(
        "https://dam-api.bfs.admin.ch/hub/api/dam/assets/36346663/master",
        timeout=60,
        follow_redirects=True,
    )
    r.raise_for_status()
    with open(path, "wb") as f:
        f.write(r.content)
    print(f"  Saved {len(r.content):,} bytes to {path}")
    return path


def parse_employment_excel(path):
    """Parse 1-digit ISCO employment from BFS Excel, latest year."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Jahreswerte"]

    # Find latest year column (row 5 has year headers)
    latest_col = None
    latest_year = 0
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=5, column=col).value
        if val and str(val).strip().isdigit():
            year = int(str(val).strip())
            if year > latest_year:
                latest_year = year
                latest_col = col

    print(f"  Latest year in Excel: {latest_year}")

    # Parse major group employment (first section before gender breakdown)
    major_employment = {}
    sub_labels = {"Schweizer", "Ausländer", "Deutschland", "Frankreich", "Italien",
                  "Österreich", "Spanien", "Portugal", "Übrige", "EU/EFTA", "Drittstaaten"}
    seen_first_total = False

    for row in range(5, 200):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a is None:
            continue
        name = str(cell_a).strip()

        if name == "Total":
            if seen_first_total:
                break  # Second total = gender breakdown starts
            seen_first_total = True
            continue

        if any(name.startswith(s) for s in sub_labels):
            continue

        if name in BFS_MAJOR_GROUP_NAMES:
            emp_val = ws.cell(row=row, column=latest_col).value
            if emp_val and str(emp_val) != "X":
                code = BFS_MAJOR_GROUP_NAMES[name]
                major_employment[code] = float(emp_val) * 1000  # Convert from thousands
                print(f"    ISCO {code}: {name} = {float(emp_val):.1f}k")

    return major_employment, latest_year


def distribute_to_subgroups(major_employment):
    """Distribute 1-digit employment to 2-digit sub-groups using weights."""
    employment = {}
    for major, total in major_employment.items():
        weights = SUBGROUP_WEIGHTS.get(major, {})
        for code, weight in weights.items():
            employment[code] = round(total * weight)
    return employment


def fetch_wages_from_api():
    """Fetch real median wages by ISCO 2-digit from BFS STAT-TAB API."""
    print("  Querying BFS STAT-TAB API for ESS 2022 wage data...")

    url = "https://www.pxweb.bfs.admin.ch/api/v1/de/px-x-0304010000_205/px-x-0304010000_205.px"

    # All 2-digit ISCO codes available in the wage table
    isco_codes = [
        "11", "12", "13", "14", "21", "22", "23", "24", "25", "26",
        "31", "32", "33", "34", "35", "41", "42", "43", "44",
        "51", "52", "53", "54", "61", "62", "71", "72", "73", "74", "75",
        "81", "82", "83", "91", "92", "93", "94", "96",
    ]

    query = {
        "query": [
            {"code": "Jahr", "selection": {"filter": "item", "values": ["2022"]}},
            {"code": "Grossregion", "selection": {"filter": "item", "values": ["-1"]}},
            {"code": "Berufsgruppe", "selection": {"filter": "item", "values": isco_codes}},
            {"code": "Lebensalter", "selection": {"filter": "item", "values": ["-1"]}},
            {"code": "Geschlecht", "selection": {"filter": "item", "values": ["-1"]}},
            {"code": "Zentralwert und andere Perzentile", "selection": {"filter": "item", "values": ["1"]}},
        ],
        "response": {"format": "json-stat2"},
    }

    r = httpx.post(url, json=query, timeout=30)
    r.raise_for_status()
    data = r.json()

    # Parse json-stat2 format
    occ_dim = data["dimension"]["Berufsgruppe"]["category"]
    occ_index = occ_dim["index"]
    values = data["value"]

    wages = {}
    for code, idx in occ_index.items():
        if idx < len(values) and values[idx] is not None:
            wages[code] = round(values[idx])

    print(f"  Fetched wages for {len(wages)} occupation groups")
    return wages


def build_occupations():
    """Build the complete occupation dataset from real BFS data."""
    print("\n--- Employment Data (BFS SAKE) ---")
    excel_path = download_employment_excel()
    major_employment, emp_year = parse_employment_excel(excel_path)
    employment = distribute_to_subgroups(major_employment)

    print(f"\n--- Wage Data (BFS ESS 2022) ---")
    wages = fetch_wages_from_api()

    print(f"\n--- Building occupation dataset ---")
    occupations = []
    for code, info in sorted(ISCO_2DIGIT.items()):
        emp = employment.get(code, 0)
        wage_monthly = wages.get(code, 0)

        # Skip groups with very low estimated employment
        if emp < 5000:
            print(f"  Skipping {code} ({info['title']}): {emp:,} employed")
            continue

        occupations.append({
            "code": code,
            "title": info["title"],
            "major_group": info["major"],
            "major_group_name": info["major_name"],
            "employment": emp,
            "wage_monthly": wage_monthly,
            "wage_annual": wage_monthly * 12,
            "skill_level": info["skill_level"],
            "education": SKILL_LEVEL_LABELS[info["skill_level"]],
        })

    return occupations, emp_year


def main():
    print("=" * 60)
    print("Swiss Jobs: Fetching occupation data from BFS")
    print("=" * 60)

    occupations, emp_year = build_occupations()

    with open("data/occupations.json", "w") as f:
        json.dump(occupations, f, indent=2, ensure_ascii=False)

    total_emp = sum(o["employment"] for o in occupations)
    print(f"\nSaved {len(occupations)} occupations to data/occupations.json")
    print(f"Total employment: {total_emp:,} (SAKE {emp_year})")
    print(f"Wage range: CHF {min(o['wage_monthly'] for o in occupations):,} - "
          f"{max(o['wage_monthly'] for o in occupations):,}/month (ESS 2022)")

    print(f"\n{'Major Group':<25} {'Count':>6} {'Employment':>12}")
    print("-" * 50)
    groups = {}
    for o in occupations:
        g = o["major_group_name"]
        groups.setdefault(g, {"count": 0, "emp": 0})
        groups[g]["count"] += 1
        groups[g]["emp"] += o["employment"]
    for g, v in sorted(groups.items(), key=lambda x: -x[1]["emp"]):
        print(f"{g:<25} {v['count']:>6} {v['emp']:>12,}")


if __name__ == "__main__":
    main()
